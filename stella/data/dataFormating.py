import pandas as pd
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import tkinter.font as tkFont
import ttkbootstrap as btk
from ttkbootstrap.constants import *
import openpyxl
import xlsxwriter
import datetime
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import (
    Paragraph,
    ParagraphProperties,
    CharacterProperties,
    Font,
)
import os
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
import matplotlib.dates as mdates
from ..utility.message import Message


class DataFormating:
    def __init__(self, root, data, option, name):
        # File save
        self.data = data
        self.option = option
        try:
            # Message during data processing:
            self.popup_win_2 = tk.Toplevel(root)
            width = 205
            height = 64
            screenwidth = self.popup_win_2.winfo_screenwidth()
            screenheight = self.popup_win_2.winfo_screenheight()
            alignstr = "%dx%d+%d+%d" % (
                width,
                height,
                (screenwidth - width) / 2,
                (screenheight - height) / 6,
            )
            self.popup_win_2.geometry(alignstr)
            self.popup_win_2.resizable(width=False, height=False)
            self.popup_win_2.attributes("-topmost", True)
            self.popup_win_2.minsize(205, 64)
            self.popup_win_2.maxsize(205, 64)

            popup_label = tk.Label(self.popup_win_2)
            popup_label["anchor"] = "n"
            ft = tkFont.Font(family="Helvetica", size=15)
            popup_label["font"] = ft
            popup_label["fg"] = "#333333"
            popup_label["justify"] = "center"
            popup_label["text"] = "Save file and wait"
            popup_label.place(x=10, y=10, width=182, height=37)

            self.savepath = filedialog.asksaveasfilename(
                parent=root,
                initialfile=f"{name}_Log",
                defaultextension=".xlsx",
                filetypes=(
                    ("Excel Files", "*.xlsx"),
                    ("all files", ".*"),
                ),
            )
            writer = pd.ExcelWriter(self.savepath, engine="xlsxwriter")
            self.data.to_excel(
                writer,
                "Data",
                index=False,
                startrow=0,
                startcol=0,
            )
            pd.io.formats.excel.header_style = None
            workbook = writer.book

            workbook.close()
            self.wb = openpyxl.load_workbook(self.savepath)
            self.sheet = self.wb["Data"]
            for col in self.sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                self.sheet.column_dimensions[column].width = adjusted_width
            self.combobox(root)
        except:
            self.popup_win_2.destroy()
            Message("error", "File could not be saved", 0)
            return

    def combobox(self, root):
        # Axis X range selection:
        self.popup_win_5 = tk.Toplevel(root)
        width = 325
        height = 280
        screenwidth = self.popup_win_5.winfo_screenwidth()
        screenheight = self.popup_win_5.winfo_screenheight()
        alignstr = "%dx%d+%d+%d" % (
            width,
            height,
            (screenwidth - width) / 2,
            (screenheight - height) / 2,
        )
        self.popup_win_5.geometry(alignstr)
        self.popup_win_5.resizable(width=False, height=False)
        self.popup_win_5.attributes("-topmost", True)
        self.popup_win_5.minsize(325, 280)
        self.popup_win_5.maxsize(325, 280)

        def chart_viewer(self):
            # Chart preview with data selection:
            try:
                if self.chart_view_var.get() == 1:
                    chart_list = pd.read_excel(self.savepath)
                    chart_list = pd.DataFrame(chart_list)
                    self.popup_win_6 = tk.Toplevel(self.popup_win_5)
                    width = 700
                    height = 770
                    screenwidth = self.popup_win_6.winfo_screenwidth()
                    screenheight = self.popup_win_6.winfo_screenheight()
                    alignstr = "%dx%d-%d+%d" % (
                        width,
                        height,
                        (screenwidth - width) / 16,
                        (screenheight - height) / 2,
                    )
                    self.popup_win_6.geometry(alignstr)
                    self.popup_win_6.resizable(width=False, height=False)
                    self.popup_win_6.attributes("-topmost", True)
                    self.popup_win_6.minsize(700, 770)
                    self.popup_win_6.maxsize(700, 770)

                    chartFrame = tk.Frame(self.popup_win_6)
                    chartFrame.place(x=0, y=0, width=700, height=720)

                    button_close = btk.Button(
                        self.popup_win_6,
                        text="Close",
                        command=lambda: [refresh(self), self.popup_win_6.destroy()],
                        bootstyle=DANGER,
                    )
                    button_close.place(x=240, y=685, width=82, height=30)

                    button_refresh = btk.Button(
                        self.popup_win_6,
                        text="Refresh",
                        command=lambda: refresh(self),
                        bootstyle=SUCCESS,
                    )
                    button_refresh.place(x=340, y=685, width=82, height=30)

                    label_select = tk.Label(self.popup_win_6)
                    label_select["anchor"] = "n"
                    ft = tkFont.Font(family="Helvetica", size=11)
                    label_select["font"] = ft
                    label_select["fg"] = "#333333"
                    label_select["justify"] = "left"
                    label_select["text"] = (
                        "Press ',' button to select min. range (axis X and Y).\nPress '.' button to select max. range (axis X and Y)."
                    )
                    label_select.place(x=0, y=720, width=400, height=60)

                    def refresh(self):
                        # Combobox refresh
                        self.cb1.delete(0, tk.END)
                        self.cb1.insert(tk.END, self.combo_start)
                        self.cb2.delete(0, tk.END)
                        self.cb2.insert(tk.END, self.combo_end)
                        self.value_min.delete(0, tk.END)
                        self.value_min.insert(tk.END, self.combo_min)
                        self.value_max.delete(0, tk.END)
                        self.value_max.insert(tk.END, self.combo_max)

                    figure = Figure(figsize=(6, 4), dpi=100)
                    figure_canvas = FigureCanvasTkAgg(figure, chartFrame)
                    NavigationToolbar2Tk(figure_canvas, chartFrame)

                    colors_chambers = [
                        "#752524",
                        "#F80E0A",
                        "#173B96",
                        "#1F95F8",
                        "#2AA31E",
                        "#000000",
                    ]  # Climatic and TS chambers.
                    colors_hum_bath = [
                        "#752524",
                        "#F80E0A",
                        "#0A6E44",
                        "#2AA31E",
                        "#000000",
                    ]  # Salt and splash.
                    colors_salt_hum = [
                        "#752524",
                        "#F80E0A",
                        "#0A6E44",
                        "#2AA31E",
                        "#173B96",
                        "#1F95F8",
                    ]  # Salt humidity
                    colors_gas_chamber = [
                        "#752524",
                        "#F80E0A",
                        "#F49D1E",
                        "#173B96",
                        "#1F95F8",
                        "#0AB1C8",
                    ]
                    ax = figure.subplots()
                    if self.option["hum"] == True:
                        ax.set_ylabel("Temperature[°C] / Humidity[%rH]")
                    else:
                        ax.set_ylabel("Temperature [°C]")
                    if (
                        self.option["samples"] == True
                        or self.option["timeabsOne"] == True
                    ):
                        for i in range(2, int(len(chart_list.columns))):
                            chart_list.plot(
                                kind="line", x=0, y=int(i), ax=ax, x_compat=True
                            )
                    elif self.option["timeabsAll"] == True:
                        try:
                            data = 1
                            for col in chart_list:
                                chart_list.plot(
                                    kind="line", x=0, y=int(data), ax=ax, x_compat=True
                                )
                                data += 2
                        except:
                            pass
                    elif self.option["indigo"] == True:
                        try:
                            for i in range(1, 3):
                                chart_list.plot(
                                    kind="line", x=0, y=i, ax=ax, x_compat=True
                                )
                        except:
                            pass
                        try:
                            for i in range(4, 6):
                                chart_list.plot(
                                    kind="line", x=3, y=i, ax=ax, x_compat=True
                                )
                        except:
                            pass
                    elif self.option["grafana"] == True:
                        try:
                            for i in range(1, 3):
                                chart_list.plot(
                                    kind="line", x=0, y=i, ax=ax, x_compat=True
                                )
                        except:
                            pass
                        try:
                            for i in range(4, 6):
                                chart_list.plot(
                                    kind="line", x=0, y=i, ax=ax, x_compat=True
                                )
                        except:
                            pass
                    elif self.option["rotronic"] == True:
                        time = 0
                        hums = 1
                        temp = 2
                        try:
                            for col in chart_list:
                                chart_list.plot(
                                    kind="line", x=time, y=hums, ax=ax, x_compat=True
                                )
                                chart_list.plot(
                                    kind="line", x=time, y=temp, ax=ax, x_compat=True
                                )
                                time += 3
                                hums += 3
                                temp += 3
                        except:
                            pass
                    elif self.option["keithleymanager"] == True:
                        time = 0
                        temp = 1
                        try:
                            for col in chart_list:
                                chart_list.plot(
                                    kind="line", x=time, y=temp, ax=ax, x_compat=True
                                )
                                time += 2
                                temp += 2
                        except:
                            pass
                    elif self.option["gasChamber"] == True:
                        x = 0
                        for i in range(1, 7):
                            chart_list.plot(
                                kind="line",
                                x=0,
                                y=i,
                                color=colors_gas_chamber[x],
                                ax=ax,
                                x_compat=True,
                            )
                            x += 1
                    elif (
                        self.option["humBath"] == True
                        and self.option["indexZUT"] == True
                    ):
                        x = 0
                        for i in range(1, 5):
                            chart_list.plot(
                                kind="line",
                                x=0,
                                y=i,
                                color=colors_hum_bath[x],
                                ax=ax,
                                x_compat=True,
                            )
                            x += 1
                    elif (
                        self.option["humBath"] == True
                        and self.option["indexZUT"] == False
                    ):
                        x = 0
                        for i in range(1, int(len(chart_list.columns))):
                            chart_list.plot(
                                kind="line",
                                x=0,
                                y=i,
                                color=colors_hum_bath[x],
                                ax=ax,
                                x_compat=True,
                            )
                            x += 1
                    elif (
                        self.option["humBath"] == False
                        and self.option["indexZUT"] == True
                    ):
                        x = 0
                        for i in range(1, 5):
                            chart_list.plot(
                                kind="line",
                                x=0,
                                y=i,
                                color=colors_chambers[x],
                                ax=ax,
                                x_compat=True,
                            )
                            x += 1
                    elif self.option["coolingSys"] == True:
                        x = 0
                        for i in range(1, 4):
                            chart_list.plot(
                                kind="line",
                                x=0,
                                y=i,
                                color=colors_salt_hum[x],
                                ax=ax,
                                x_compat=True,
                            )
                            x += 1
                    elif self.option["saltHum"] == True:
                        x = 0
                        for i in range(1, int(len(chart_list.columns))):
                            chart_list.plot(
                                kind="line",
                                x=0,
                                y=i,
                                color=colors_salt_hum[x],
                                ax=ax,
                                x_compat=True,
                            )
                            x += 1
                    elif self.option["agilent"] == True:
                        x = 0
                        for i in range(1, int(len(chart_list.columns))):
                            chart_list.plot(kind="line", x=0, y=i, ax=ax, x_compat=True)
                            x += 1
                    elif self.option["insight"] == True:
                        if int(len(chart_list.columns)) == 2:
                            chart_list.plot(
                                kind="line",
                                x=0,
                                y=1,
                                color="#FFA500",
                                ax=ax,
                                x_compat=True,
                            )
                        else:
                            x = 0
                            for i in range(1, int(len(chart_list.columns))):
                                chart_list.plot(
                                    kind="line", x=0, y=i, ax=ax, x_compat=True
                                )
                                x += 1
                    elif self.option["secasi"] == True:
                        x = 0
                        for i in range(1, int(len(chart_list.columns))):
                            chart_list.plot(
                                kind="line",
                                x=0,
                                y=int(i),
                                color=colors_chambers[x],
                                ax=ax,
                                x_compat=True,
                            )
                            x += 1
                    else:
                        try:
                            x = 0
                            for i in range(1, int(len(chart_list.columns))):
                                chart_list.plot(
                                    kind="line",
                                    x=0,
                                    y=i,
                                    color=colors_chambers[x],
                                    ax=ax,
                                    x_compat=True,
                                )
                                x += 1
                        except:
                            pass
                    pos = ax.get_position()
                    ax.set_position([pos.x0, pos.y0, pos.width, pos.height * 0.98])
                    ax.legend(
                        loc="upper center",
                        fancybox=True,
                        fontsize="small",
                        bbox_to_anchor=(0.5, 1.15),
                        ncol=2,
                    )
                    if (
                        self.option["samples"] == True
                        or self.option["timeabsOne"] == True
                        or self.option["secasi"] == True
                    ):
                        pass
                    else:
                        ax.xaxis.set_major_formatter(
                            mdates.DateFormatter("%y-%m-%d %H:%M")
                        )
                    ax.grid(True)
                    figure_canvas.draw()
                    figure_canvas.get_tk_widget().pack(
                        side=tk.TOP, fill=tk.BOTH, expand=1
                    )

                    def on_press(event):
                        if event.key == ",":
                            ix = event.xdata
                            iy = event.ydata
                            if (
                                self.option["samples"] == True
                                or self.option["secasi"] == True
                                or self.option["timeabsOne"] == True
                                or self.option["timeabsAll"] == True
                            ):
                                self.date_start = int(ix)
                            else:
                                try:
                                    self.date_start = str(mdates.num2date(ix))
                                    self.date_start = self.date_start.split(".")
                                    self.date_start = self.date_start[0]
                                except:
                                    self.date_start = datetime.datetime.fromtimestamp(
                                        ix
                                    ).strftime("%Y-%m-%d %H:%M:%S")
                            self.combo_min = int(iy)
                            self.combo_start = self.date_start

                        elif event.key == ".":
                            ix = event.xdata
                            iy = event.ydata
                            if (
                                self.option["samples"] == True
                                or self.option["secasi"] == True
                                or self.option["timeabsOne"] == True
                                or self.option["timeabsAll"] == True
                            ):
                                self.date_end = int(ix)
                            else:
                                try:
                                    self.date_end = str(mdates.num2date(ix))
                                    self.date_end = self.date_end.split(".")
                                    self.date_end = self.date_end[0]
                                except:
                                    self.date_end = datetime.datetime.fromtimestamp(
                                        ix
                                    ).strftime("%Y-%m-%d %H:%M:%S")
                            self.combo_max = int(iy)
                            self.combo_end = self.date_end

                    figure.canvas.mpl_connect("key_press_event", on_press)

                    # Cross-hair
                    value = chart_list.loc[1, :].values[0]
                    if (
                        self.option["samples"] == True
                        or self.option["secasi"] == True
                        or self.option["timeabsOne"] == True
                        or self.option["timeabsAll"] == True
                    ):
                        pass
                    else:
                        value = mdates.date2num(value)
                    line1 = ax.axvline(x=value, color="r", lw=0.8, ls="--")
                    line2 = ax.axhline(y=0, color="r", lw=0.8, ls="--")
                    figure_canvas.draw()

                    def onMouseMove(event):
                        ix = event.xdata
                        iy = event.ydata
                        if iy == None or ix == None:
                            pass
                        else:
                            line1.set_xdata([ix])
                            line2.set_ydata([iy])
                            figure_canvas.draw()

                    figure.canvas.mpl_connect("motion_notify_event", onMouseMove)

                else:
                    self.popup_win_6.destroy()
            except:
                try:
                    self.popup_win_2.destroy()
                    self.popup_win_5.destroy()
                    self.popup_win_6.destroy()
                except:
                    pass
                Message(
                    "error",
                    "Something went wrong"
                    "\nPlease contact with: piotr.klys92@gmail.com",
                    0,
                )

        if self.option["secasi"] or self.option["samples"] == True:
            data_list = []
            for data in self.sheet.iter_rows(
                min_row=2, max_col=1, max_row=self.sheet.max_row, values_only=True
            ):
                data_list.append(data)
            data_list = pd.DataFrame(data_list)
            data_list = data_list[0].tolist()
        else:
            if self.option["timeabsAll"] == True or self.option["timeabsOne"] == True:
                data_list = []
                for data in self.sheet.iter_rows(
                    min_row=2, max_col=1, max_row=self.sheet.max_row, values_only=True
                ):
                    data_list.append(data)
                data_list = pd.DataFrame(data_list)
                data_list[0] = data_list[0].astype(str)
                data_list = data_list[0].tolist()
            else:
                data_list = []
                for data in self.sheet.iter_rows(
                    min_row=2, max_col=1, max_row=self.sheet.max_row, values_only=True
                ):
                    data_list.append(data)
                data_list = pd.DataFrame(data_list)
                data_list[0] = pd.to_datetime(
                    data_list[0], format="%y-%m-%d %H:%M:%S", errors="coerce"
                )
                data_list[0] = data_list[0].astype(str)
                data_list = data_list[0].tolist()

        if (
            self.option["samples"] == True
            or self.option["secasi"] == True
            or self.option["timeabsOne"] == True
            or self.option["timeabsAll"] == True
        ):

            def cb1_value(event):
                self.min_value = self.cb1_var.get()
                data_list2 = [
                    int(i) for i in data_list if int(i) > int(self.cb1_var.get())
                ]
                data_list2.sort()
                self.cb2.config(values=data_list2)

        else:

            def cb1_value(event):
                self.min_value = self.cb1_var.get()
                data_list2 = [i for i in data_list if i > self.cb1_var.get()]
                data_list2.sort()
                self.cb2.config(values=data_list2)

        self.cb1_var = tk.StringVar()
        self.cb1 = ttk.Combobox(
            self.popup_win_5, values=data_list, textvariable=self.cb1_var, width=20
        )
        self.cb1.place(x=80, y=100, width=189, height=30)
        self.cb1.bind("<<ComboboxSelected>>", cb1_value)

        self.cb2_var = tk.StringVar()
        self.cb2 = ttk.Combobox(self.popup_win_5, textvariable=self.cb2_var, width=20)
        self.cb2.place(x=80, y=135, width=189, height=30)
        try:
            button_new_value = btk.Button(
                self.popup_win_5,
                text="Selected",
                command=lambda: [
                    self.popup_win_5.destroy(),
                    print_values(
                        self,
                        self.cb1_var.get(),
                        self.cb2_var.get(),
                        self.min_variable.get(),
                        self.max_variable.get(),
                        self.popup_win_6.destroy(),
                    ),
                ],
                bootstyle=DARK,
            )
            button_new_value.place(x=20, y=50, width=82, height=30)

            button_old_value = btk.Button(
                self.popup_win_5,
                text="Total",
                command=lambda: [
                    self.popup_win_5.destroy(),
                    old_values(self),
                    self.popup_win_6.destroy(),
                ],
                bootstyle=DARK,
            )
            button_old_value.place(x=120, y=50, width=82, height=30)

            button_all_values = btk.Button(
                self.popup_win_5,
                text="Both",
                command=lambda: [
                    self.popup_win_5.destroy(),
                    print_values(
                        self,
                        self.cb1_var.get(),
                        self.cb2_var.get(),
                        self.min_variable.get(),
                        self.max_variable.get(),
                        both=True,
                    ),
                    old_values(self),
                    self.popup_win_6.destroy(),
                ],
                bootstyle=DARK,
            )
            button_all_values.place(x=220, y=50, width=82, height=30)
        except:
            pass

        label_start = tk.Label(self.popup_win_5)
        label_start["anchor"] = "n"
        ft = tkFont.Font(family="Helvetica", size=11)
        label_start["font"] = ft
        label_start["fg"] = "#333333"
        label_start["justify"] = "center"
        label_start["text"] = "From:"
        label_start.place(x=10, y=105, width=50, height=30)

        label_end = tk.Label(self.popup_win_5)
        label_end["anchor"] = "n"
        ft = tkFont.Font(family="Helvetica", size=11)
        label_end["font"] = ft
        label_end["fg"] = "#333333"
        label_end["justify"] = "center"
        label_end["text"] = "To:"
        label_end.place(x=10, y=140, width=50, height=30)

        label_data = tk.Label(self.popup_win_5)
        label_data["anchor"] = "n"
        ft = tkFont.Font(family="Helvetica", size=15)
        label_data["font"] = ft
        label_data["fg"] = "#333333"
        label_data["justify"] = "center"
        label_data["text"] = "Axis X - scale"
        label_data.place(x=25, y=10, width=280, height=36)

        button_close = btk.Button(
            self.popup_win_5,
            text="Close",
            command=lambda: [
                self.popup_win_5.destroy(),
                self.popup_win_6.destroy(),
            ],
            bootstyle=DANGER,
        )
        button_close.place(x=120, y=240, width=82, height=30)

        self.chart_view_var = tk.IntVar()
        self.chart_value = self.chart_view_var.get()
        chart_view = tk.Checkbutton(
            self.popup_win_5,
            text="Chart preview",
            variable=self.chart_view_var,
            command=lambda: chart_viewer(self),
        )
        ft = tkFont.Font(family="Helvetica", size=9)
        chart_view["font"] = ft
        chart_view["fg"] = "#333333"
        chart_view["justify"] = "center"
        chart_view["offvalue"] = 0
        chart_view["onvalue"] = 1
        chart_view.place(x=10, y=240, width=95, height=25)

        label_min = tk.Label(self.popup_win_5)
        label_min["anchor"] = "n"
        ft = tkFont.Font(family="Helvetica", size=11)
        label_min["font"] = ft
        label_min["fg"] = "#333333"
        label_min["justify"] = "center"
        label_min["text"] = "Min:"
        label_min.place(x=20, y=180, width=50, height=30)

        label_scale = tk.Label(self.popup_win_5)
        label_scale["anchor"] = "n"
        ft = tkFont.Font(family="Helvetica", size=9)
        label_scale["font"] = ft
        label_scale["fg"] = "#333333"
        label_scale["justify"] = "center"
        label_scale["text"] = "Axis Y Scale:"
        label_scale.place(x=120, y=180, width=80, height=30)

        label_max = tk.Label(self.popup_win_5)
        label_max["anchor"] = "n"
        ft = tkFont.Font(family="Helvetica", size=11)
        label_max["font"] = ft
        label_max["fg"] = "#333333"
        label_max["justify"] = "center"
        label_max["text"] = "Max:"
        label_max.place(x=220, y=180, width=50, height=30)

        self.min_variable = tk.StringVar()
        self.value_min = tk.Entry(
            self.popup_win_5, justify="center", textvariable=self.min_variable
        )
        self.value_min["bg"] = "#ffffff"
        ft = tkFont.Font(family="Helvetica", size=11)
        self.value_min["font"] = ft
        self.value_min["fg"] = "#333333"
        self.value_min.place(x=20, y=200, width=82, height=30)

        self.ax_scale = tk.StringVar()
        self.scale_ax = tk.Entry(
            self.popup_win_5, justify="center", textvariable=self.ax_scale
        )
        self.scale_ax["bg"] = "#ffffff"
        ft = tkFont.Font(family="Helvetica", size=11)
        self.scale_ax["font"] = ft
        self.scale_ax["fg"] = "#333333"
        self.scale_ax.insert(0, "5")
        self.scale_ax.place(x=120, y=200, width=82, height=30)

        self.max_variable = tk.StringVar()
        self.value_max = tk.Entry(
            self.popup_win_5, justify="center", textvariable=self.max_variable
        )
        self.value_max["bg"] = "#ffffff"
        ft = tkFont.Font(family="Helvetica", size=11)
        self.value_max["font"] = ft
        self.value_max["fg"] = "#333333"
        self.value_max.place(x=220, y=200, width=82, height=30)

        def print_values(
            self, combo_start_val, combo_end_val, combo_min_val, combo_max_val, both
        ):
            try:
                if self.option["hum"] == True:
                    self.sheet_chart = self.wb.create_chartsheet("Chart")
                    self.chart = ScatterChart()
                    self.chart.x_axis.number_format = "yyyy-mm-dd HH:MM:SS"
                    self.chart.x_axis.majorTimeUnit = "months"
                    self.chart.x_axis.tickLblPos = "low"
                    self.chart.legend.position = "b"
                    self.chart.x_axis.title = "Date and Time [yyyy-mm-dd hh:mm:ss]"
                    self.chart.y_axis.title = "Temperature[°C] / Humidity[%rH]"

                    # Text format - Axis
                    normal_font = Font(typeface="Calibri")
                    cp_text = CharacterProperties(latin=normal_font, sz=900, b=False)
                    self.chart.x_axis.txPr = RichText(
                        p=[
                            Paragraph(
                                pPr=ParagraphProperties(defRPr=cp_text),
                                endParaRPr=cp_text,
                            )
                        ]
                    )
                    self.chart.y_axis.txPr = RichText(
                        p=[
                            Paragraph(
                                pPr=ParagraphProperties(defRPr=cp_text),
                                endParaRPr=cp_text,
                            )
                        ]
                    )

                    # Text format - Title
                    cp_text1 = CharacterProperties(latin=normal_font, sz=1000, b=False)
                    pp1 = ParagraphProperties(defRPr=cp_text1)
                    self.chart.x_axis.title.tx.rich.p[0].pPr = pp1
                    self.chart.y_axis.title.tx.rich.p[0].pPr = pp1
                    self.chart.x_axis.txPr.properties.rot = "-2700000"
                    self.chart.y_axis.majorUnit = self.ax_scale.get()
                    self.min_value = combo_start_val
                    self.max_value = combo_end_val
                    try:
                        self.chart.y_axis.scaling.min = combo_min_val
                        self.chart.y_axis.scaling.max = combo_max_val
                    except:
                        pass
                    min_scale = pd.Timestamp(self.min_value)
                    max_scale = pd.Timestamp(self.max_value)
                    min_converted = (
                        min_scale - pd.Timestamp("1899-12-30")
                    ).total_seconds() / 86400
                    max_converted = (
                        max_scale - pd.Timestamp("1899-12-30")
                    ).total_seconds() / 86400
                    self.chart.x_axis.scaling.min = min_converted
                    self.chart.x_axis.scaling.max = max_converted

                    if self.option["rotronic"] == True:
                        col_time = self.sheet.min_column
                        col_meas_range_min = 2
                        col_meas_range_max = 4
                        while col_time <= self.sheet.max_column:
                            data = Reference(
                                self.sheet,
                                min_col=col_time,
                                min_row=2,
                                max_row=self.sheet.max_row,
                            )
                            col_time += 3
                            for i in range(col_meas_range_min, col_meas_range_max):
                                values = Reference(
                                    self.sheet,
                                    min_col=i,
                                    min_row=1,
                                    max_row=self.sheet.max_row,
                                )
                                series = Series(values, data, title_from_data=True)
                                self.chart.series.append(series)
                            col_meas_range_min += 3
                            col_meas_range_max += 3
                    elif self.option["indigo"] == True:
                        data = Reference(
                            self.sheet, min_col=1, min_row=2, max_row=self.sheet.max_row
                        )
                        for i in range(2, 4):
                            values = Reference(
                                self.sheet,
                                min_col=i,
                                min_row=1,
                                max_row=self.sheet.max_row,
                            )
                            series = Series(values, data, title_from_data=True)
                            self.chart.series.append(series)
                        try:
                            data2 = Reference(
                                self.sheet,
                                min_col=4,
                                min_row=2,
                                max_row=self.sheet.max_row,
                            )
                            for i in range(5, 7):
                                values2 = Reference(
                                    self.sheet,
                                    min_col=i,
                                    min_row=1,
                                    max_row=self.sheet.max_row,
                                )
                                series2 = Series(values2, data2, title_from_data=True)
                                self.chart.series.append(series2)
                        except:
                            pass
                    else:
                        if self.option["indexZUT"] == True:
                            index = int(6)
                        elif self.option["coolingSys"] == True:
                            index = int(5)
                        else:
                            index = int(self.sheet.max_column + 1)
                        data = Reference(
                            self.sheet, min_col=1, min_row=2, max_row=self.sheet.max_row
                        )
                        for i in range(2, index):
                            values = Reference(
                                self.sheet,
                                min_col=i,
                                min_row=1,
                                max_row=self.sheet.max_row,
                            )
                            series = Series(values, data, title_from_data=True)
                            self.chart.series.append(series)
                    self.sheet_chart.add_chart(self.chart)
                    for serie in self.chart.series:
                        serie.graphicalProperties.line.width = 15000
                    if (
                        self.option["grafana"] == True
                        or self.option["rotronic"] == True
                        or self.option["indigo"] == True
                    ):
                        pass
                    else:
                        if self.option["insight"] == True:
                            # Temperature
                            s1 = self.chart.series[0]
                            s1.graphicalProperties.line.solidFill = "F80E0A"

                            # Humidity
                            s2 = self.chart.series[0]
                            s2.graphicalProperties.line.solidFill = "1F95F8"
                        else:
                            # Temperature Set
                            s1 = self.chart.series[0]
                            s1.graphicalProperties.line.solidFill = "752524"

                            # Temperature
                            s2 = self.chart.series[1]
                            s2.graphicalProperties.line.solidFill = "F80E0A"

                            if self.option["gasChamber"] == True:
                                # Temperature Set
                                s1 = self.chart.series[0]
                                s1.graphicalProperties.line.solidFill = "752524"

                                # Temperature IN
                                s2 = self.chart.series[1]
                                s2.graphicalProperties.line.solidFill = "F80E0A"

                                # Temperature OUT
                                s3 = self.chart.series[2]
                                s3.graphicalProperties.line.solidFill = "F49D1E"

                                # Humidity Set
                                s4 = self.chart.series[3]
                                s4.graphicalProperties.line.solidFill = "173B96"

                                # Humidity IN
                                s5 = self.chart.series[4]
                                s5.graphicalProperties.line.solidFill = "1F95F8"

                                # Humidity OUT
                                s6 = self.chart.series[5]
                                s6.graphicalProperties.line.solidFill = "0AB1C8"
                            else:
                                if self.option["saltHum"] == True:
                                    # Humidifier Set
                                    s3 = self.chart.series[2]
                                    s3.graphicalProperties.line.solidFill = "0A6E44"

                                    # Humidifier
                                    s4 = self.chart.series[3]
                                    s4.graphicalProperties.line.solidFill = "2AA31E"

                                    # Humidity Set
                                    s5 = self.chart.series[4]
                                    s5.graphicalProperties.line.solidFill = "173B96"

                                    # Humidity
                                    s6 = self.chart.series[5]
                                    s6.graphicalProperties.line.solidFill = "1F95F8"
                                else:
                                    try:
                                        # Humidity Set
                                        s3 = self.chart.series[2]
                                        s3.graphicalProperties.line.solidFill = "173B96"
                                    except:
                                        pass

                                    try:
                                        # Humidity
                                        s4 = self.chart.series[3]
                                        s4.graphicalProperties.line.solidFill = "1F95F8"
                                    except:
                                        pass

                                    try:
                                        # Temperature Bath
                                        s5 = self.chart.series[4]
                                        s5.graphicalProperties.line.solidFill = "2AA31E"
                                    except:
                                        pass

                    self.wb.save(self.savepath)
                    self.popup_win_2.destroy()
                    if both == True:
                        pass
                    else:
                        try:
                            os.startfile(self.savepath)
                        except:
                            Message(
                                "warning",
                                "File can not be opened by Stella"
                                "\nPlease open it manually",
                                0,
                            )
                else:
                    self.sheet_chart = self.wb.create_chartsheet("Chart")
                    self.chart = ScatterChart()
                    if self.option["timeabsOne"] == True:
                        self.chart.x_axis.tickLblPos = "low"
                        self.chart.legend.position = "b"

                        self.chart.x_axis.title = "Samples"
                        self.chart.y_axis.title = "Temperature[°C]"

                        # Text format - Axis
                        normal_font = Font(typeface="Calibri")
                        cp_text = CharacterProperties(
                            latin=normal_font, sz=900, b=False
                        )
                        self.chart.x_axis.txPr = RichText(
                            p=[
                                Paragraph(
                                    pPr=ParagraphProperties(defRPr=cp_text),
                                    endParaRPr=cp_text,
                                )
                            ]
                        )
                        self.chart.y_axis.txPr = RichText(
                            p=[
                                Paragraph(
                                    pPr=ParagraphProperties(defRPr=cp_text),
                                    endParaRPr=cp_text,
                                )
                            ]
                        )

                        # Text format - Title
                        cp_text1 = CharacterProperties(
                            latin=normal_font, sz=1000, b=False
                        )
                        pp1 = ParagraphProperties(defRPr=cp_text1)
                        self.chart.x_axis.title.tx.rich.p[0].pPr = pp1
                        self.chart.y_axis.title.tx.rich.p[0].pPr = pp1
                        self.chart.x_axis.txPr.properties.rot = "-2700000"
                        self.chart.y_axis.majorUnit = self.ax_scale.get()
                        self.min_value = combo_start_val
                        self.max_value = combo_end_val
                        try:
                            self.chart.y_axis.scaling.min = combo_min_val
                            self.chart.y_axis.scaling.max = combo_max_val
                        except:
                            pass
                        self.chart.x_axis.scaling.min = self.min_value
                        self.chart.x_axis.scaling.max = self.max_value

                        index = int(self.sheet.max_column + 1)

                        data = Reference(
                            self.sheet, min_col=2, min_row=2, max_row=self.sheet.max_row
                        )
                        for i in range(3, index):
                            values = Reference(
                                self.sheet,
                                min_col=i,
                                min_row=1,
                                max_row=self.sheet.max_row,
                            )
                            series = Series(values, data, title_from_data=True)
                            self.chart.series.append(series)
                        self.sheet_chart.add_chart(self.chart)
                        for serie in self.chart.series:
                            serie.graphicalProperties.line.width = 15000
                    else:
                        if self.option["samples"] == True:
                            self.chart.x_axis.tickLblPos = "low"
                            self.chart.legend.position = "b"

                            self.chart.x_axis.title = "Samples"
                            self.chart.y_axis.title = "Temperature[°C]"

                            # Text format - Axis
                            normal_font = Font(typeface="Calibri")
                            cp_text = CharacterProperties(
                                latin=normal_font, sz=900, b=False
                            )
                            self.chart.x_axis.txPr = RichText(
                                p=[
                                    Paragraph(
                                        pPr=ParagraphProperties(defRPr=cp_text),
                                        endParaRPr=cp_text,
                                    )
                                ]
                            )
                            self.chart.y_axis.txPr = RichText(
                                p=[
                                    Paragraph(
                                        pPr=ParagraphProperties(defRPr=cp_text),
                                        endParaRPr=cp_text,
                                    )
                                ]
                            )

                            # Text format - Title
                            cp_text1 = CharacterProperties(
                                latin=normal_font, sz=1000, b=False
                            )
                            pp1 = ParagraphProperties(defRPr=cp_text1)
                            self.chart.x_axis.title.tx.rich.p[0].pPr = pp1
                            self.chart.y_axis.title.tx.rich.p[0].pPr = pp1
                            self.chart.x_axis.txPr.properties.rot = "-2700000"
                            self.chart.y_axis.majorUnit = self.ax_scale.get()
                            self.min_value = combo_start_val
                            self.max_value = combo_end_val
                            try:
                                self.chart.y_axis.scaling.min = combo_min_val
                                self.chart.y_axis.scaling.max = combo_max_val
                            except:
                                pass
                            self.chart.x_axis.scaling.min = self.min_value
                            self.chart.x_axis.scaling.max = self.max_value

                            index = int(self.sheet.max_column + 1)

                            data = Reference(
                                self.sheet,
                                min_col=1,
                                min_row=2,
                                max_row=self.sheet.max_row,
                            )
                            for i in range(3, index):
                                values = Reference(
                                    self.sheet,
                                    min_col=i,
                                    min_row=1,
                                    max_row=self.sheet.max_row,
                                )
                                series = Series(values, data, title_from_data=True)
                                self.chart.series.append(series)
                            self.sheet_chart.add_chart(self.chart)
                            for serie in self.chart.series:
                                serie.graphicalProperties.line.width = 15000
                        else:
                            if self.option["timeabsAll"] == True:
                                self.chart.x_axis.tickLblPos = "low"
                                self.chart.legend.position = "b"

                                self.chart.x_axis.title = "Samples"
                                self.chart.y_axis.title = "Temperature[°C]"

                                # Text format - Axis
                                normal_font = Font(typeface="Calibri")
                                cp_text = CharacterProperties(
                                    latin=normal_font, sz=900, b=False
                                )
                                self.chart.x_axis.txPr = RichText(
                                    p=[
                                        Paragraph(
                                            pPr=ParagraphProperties(defRPr=cp_text),
                                            endParaRPr=cp_text,
                                        )
                                    ]
                                )
                                self.chart.y_axis.txPr = RichText(
                                    p=[
                                        Paragraph(
                                            pPr=ParagraphProperties(defRPr=cp_text),
                                            endParaRPr=cp_text,
                                        )
                                    ]
                                )

                                # Text format - Title
                                cp_text1 = CharacterProperties(
                                    latin=normal_font, sz=1000, b=False
                                )
                                pp1 = ParagraphProperties(defRPr=cp_text1)
                                self.chart.x_axis.title.tx.rich.p[0].pPr = pp1
                                self.chart.y_axis.title.tx.rich.p[0].pPr = pp1
                                self.chart.x_axis.txPr.properties.rot = "-2700000"
                                self.chart.y_axis.majorUnit = self.ax_scale.get()
                                self.min_value = combo_start_val
                                self.max_value = combo_end_val
                                try:
                                    self.chart.y_axis.scaling.min = combo_min_val
                                    self.chart.y_axis.scaling.max = combo_max_val
                                except:
                                    pass
                                self.chart.x_axis.scaling.min = self.min_value
                                self.chart.x_axis.scaling.max = self.max_value

                                col_time = 3
                                col_meas = 2
                                while col_time <= self.sheet.max_column:
                                    data = Reference(
                                        self.sheet,
                                        min_col=col_time,
                                        min_row=2,
                                        max_row=self.sheet.max_row,
                                    )
                                    col_time += 2
                                    if col_meas <= self.sheet.max_column:
                                        values = Reference(
                                            self.sheet,
                                            min_col=col_meas,
                                            min_row=1,
                                            max_row=self.sheet.max_row,
                                        )
                                        series = Series(
                                            values, data, title_from_data=True
                                        )
                                        self.chart.series.append(series)
                                    col_meas += 2
                                self.sheet_chart.add_chart(self.chart)
                                for serie in self.chart.series:
                                    serie.graphicalProperties.line.width = 15000
                            else:
                                if self.option["secasi"] == True:
                                    self.chart.x_axis.tickLblPos = "low"
                                    self.chart.legend.position = "b"

                                    self.chart.x_axis.title = "Time [s]"
                                    self.chart.y_axis.title = "Temperature[°C]"

                                    # Text format - Axis
                                    normal_font = Font(typeface="Calibri")
                                    cp_text = CharacterProperties(
                                        latin=normal_font, sz=900, b=False
                                    )
                                    self.chart.x_axis.txPr = RichText(
                                        p=[
                                            Paragraph(
                                                pPr=ParagraphProperties(defRPr=cp_text),
                                                endParaRPr=cp_text,
                                            )
                                        ]
                                    )
                                    self.chart.y_axis.txPr = RichText(
                                        p=[
                                            Paragraph(
                                                pPr=ParagraphProperties(defRPr=cp_text),
                                                endParaRPr=cp_text,
                                            )
                                        ]
                                    )

                                    # Text format - Title
                                    cp_text1 = CharacterProperties(
                                        latin=normal_font, sz=1000, b=False
                                    )
                                    pp1 = ParagraphProperties(defRPr=cp_text1)
                                    self.chart.x_axis.title.tx.rich.p[0].pPr = pp1
                                    self.chart.y_axis.title.tx.rich.p[0].pPr = pp1
                                    self.chart.x_axis.txPr.properties.rot = "-2700000"
                                    self.chart.y_axis.majorUnit = self.ax_scale.get()
                                    self.min_value = combo_start_val
                                    self.max_value = combo_end_val
                                    try:
                                        self.chart.y_axis.scaling.min = combo_min_val
                                        self.chart.y_axis.scaling.max = combo_max_val
                                    except:
                                        pass
                                    self.chart.x_axis.scaling.min = self.min_value
                                    self.chart.x_axis.scaling.max = self.max_value
                                else:
                                    self.chart.x_axis.number_format = (
                                        "yyyy-mm-dd HH:MM:SS"
                                    )
                                    self.chart.x_axis.majorTimeUnit = "months"
                                    self.chart.x_axis.tickLblPos = "low"
                                    self.chart.legend.position = "b"
                                    self.chart.x_axis.title = (
                                        "Date and Time [yyyy-mm-dd hh:mm:ss]"
                                    )
                                    self.chart.y_axis.title = "Temperature[°C]"

                                    # Text format - Axis
                                    normal_font = Font(typeface="Calibri")
                                    cp_text = CharacterProperties(
                                        latin=normal_font, sz=900, b=False
                                    )
                                    self.chart.x_axis.txPr = RichText(
                                        p=[
                                            Paragraph(
                                                pPr=ParagraphProperties(defRPr=cp_text),
                                                endParaRPr=cp_text,
                                            )
                                        ]
                                    )
                                    self.chart.y_axis.txPr = RichText(
                                        p=[
                                            Paragraph(
                                                pPr=ParagraphProperties(defRPr=cp_text),
                                                endParaRPr=cp_text,
                                            )
                                        ]
                                    )

                                    # Text format - Title
                                    cp_text1 = CharacterProperties(
                                        latin=normal_font, sz=1000, b=False
                                    )
                                    pp1 = ParagraphProperties(defRPr=cp_text1)
                                    self.chart.x_axis.title.tx.rich.p[0].pPr = pp1
                                    self.chart.y_axis.title.tx.rich.p[0].pPr = pp1
                                    self.chart.x_axis.txPr.properties.rot = "-2700000"
                                    self.chart.y_axis.majorUnit = self.ax_scale.get()
                                    self.min_value = combo_start_val
                                    self.max_value = combo_end_val
                                    try:
                                        self.chart.y_axis.scaling.min = combo_min_val
                                        self.chart.y_axis.scaling.max = combo_max_val
                                    except:
                                        pass
                                    min_scale = pd.Timestamp(self.min_value)
                                    max_scale = pd.Timestamp(self.max_value)
                                    min_converted = (
                                        min_scale - pd.Timestamp("1899-12-30")
                                    ).total_seconds() / 86400
                                    max_converted = (
                                        max_scale - pd.Timestamp("1899-12-30")
                                    ).total_seconds() / 86400
                                    self.chart.x_axis.scaling.min = min_converted
                                    self.chart.x_axis.scaling.max = max_converted
                                if self.option["keithleymanager"] == True:
                                    col_time = 1
                                    col_meas = 2
                                    while col_time <= self.sheet.max_column:
                                        data = Reference(
                                            self.sheet,
                                            min_col=col_time,
                                            min_row=2,
                                            max_row=self.sheet.max_row,
                                        )
                                        col_time += 2
                                        if col_meas <= self.sheet.max_column:
                                            values = Reference(
                                                self.sheet,
                                                min_col=col_meas,
                                                min_row=1,
                                                max_row=self.sheet.max_row,
                                            )
                                            series = Series(
                                                values, data, title_from_data=True
                                            )
                                            self.chart.series.append(series)
                                        col_meas += 2
                                    self.sheet_chart.add_chart(self.chart)
                                    for serie in self.chart.series:
                                        serie.graphicalProperties.line.width = 15000
                                else:
                                    if self.option["indexZUT"] == True:
                                        index = int(6)
                                    elif self.option["coolingSys"] == True:
                                        index = int(5)
                                    else:
                                        index = int(self.sheet.max_column + 1)
                                    data = Reference(
                                        self.sheet,
                                        min_col=1,
                                        min_row=2,
                                        max_row=self.sheet.max_row,
                                    )
                                    for i in range(2, index):
                                        values = Reference(
                                            self.sheet,
                                            min_col=i,
                                            min_row=1,
                                            max_row=self.sheet.max_row,
                                        )
                                        series = Series(
                                            values, data, title_from_data=True
                                        )
                                        self.chart.series.append(series)
                                    self.sheet_chart.add_chart(self.chart)
                                    for serie in self.chart.series:
                                        serie.graphicalProperties.line.width = 15000

                                if (
                                    self.option["agilent"] == True
                                    or self.option["keithleymanager"] == True
                                ):
                                    pass
                                else:
                                    if self.option["insight"] == True:
                                        # Temperature
                                        s1 = self.chart.series[0]
                                        s1.graphicalProperties.line.solidFill = "F80E0A"
                                    else:
                                        # Temperature Hot Set
                                        s1 = self.chart.series[0]
                                        s1.graphicalProperties.line.solidFill = "752524"

                                        # Temperature Hot
                                        s2 = self.chart.series[1]
                                        s2.graphicalProperties.line.solidFill = "F80E0A"
                                        if self.option["humBath"] == True:
                                            # Humidifier Set
                                            s3 = self.chart.series[2]
                                            s3.graphicalProperties.line.solidFill = (
                                                "0A6E44"
                                            )

                                            # Humidifier
                                            s4 = self.chart.series[3]
                                            s4.graphicalProperties.line.solidFill = (
                                                "2AA31E"
                                            )
                                        else:
                                            try:
                                                # Temperature Cold Set
                                                s3 = self.chart.series[2]
                                                s3.graphicalProperties.line.solidFill = (
                                                    "173B96"
                                                )
                                            except:
                                                pass
                                            try:
                                                # Temperature Cold
                                                s4 = self.chart.series[3]
                                                s4.graphicalProperties.line.solidFill = (
                                                    "1F95F8"
                                                )
                                            except:
                                                pass
                                            try:
                                                # Temperature Basket
                                                s5 = self.chart.series[4]
                                                s5.graphicalProperties.line.solidFill = (
                                                    "2AA31E"
                                                )
                                            except:
                                                pass

                    self.wb.save(self.savepath)
                    self.popup_win_2.destroy()
                    if both == True:
                        pass
                    else:
                        try:
                            os.startfile(self.savepath)
                        except:
                            Message(
                                "warning",
                                "File can not be opened by Stella"
                                "\nPlease open it manually",
                                0,
                            )
            except:
                try:
                    self.popup_win_2.destroy()
                except:
                    pass
                Message(
                    "error",
                    "Something went wrong"
                    "\nPlease contact with: piotr.klys92@gmail.com",
                    0,
                )

        def old_values(self):
            try:
                # Total duration chart plotting
                if self.option["hum"] == True:
                    self.sheet_chart = self.wb.create_chartsheet("ChartAll")
                    self.chart = ScatterChart()
                    self.chart.x_axis.number_format = "yyyy-mm-dd HH:MM:SS"
                    self.chart.x_axis.majorTimeUnit = "months"
                    self.chart.x_axis.tickLblPos = "low"
                    self.chart.legend.position = "b"
                    self.chart.x_axis.title = "Date and Time [yyyy-mm-dd hh:mm:ss]"
                    self.chart.y_axis.title = "Temperature[°C] / Humidity[%rH]"

                    # Text format - Axis
                    normal_font = Font(typeface="Calibri")
                    cp_text = CharacterProperties(latin=normal_font, sz=900, b=False)
                    self.chart.x_axis.txPr = RichText(
                        p=[
                            Paragraph(
                                pPr=ParagraphProperties(defRPr=cp_text),
                                endParaRPr=cp_text,
                            )
                        ]
                    )
                    self.chart.y_axis.txPr = RichText(
                        p=[
                            Paragraph(
                                pPr=ParagraphProperties(defRPr=cp_text),
                                endParaRPr=cp_text,
                            )
                        ]
                    )

                    # Text format - Title
                    cp_text1 = CharacterProperties(latin=normal_font, sz=1000, b=False)
                    pp1 = ParagraphProperties(defRPr=cp_text1)
                    self.chart.x_axis.title.tx.rich.p[0].pPr = pp1
                    self.chart.y_axis.title.tx.rich.p[0].pPr = pp1
                    self.chart.x_axis.txPr.properties.rot = "-2700000"
                    self.chart.y_axis.majorUnit = self.ax_scale.get()
                    self.min_value = self.sheet["A" + str(self.sheet.min_row + 1)].value
                    self.max_value = self.sheet["A" + str(self.sheet.max_row)].value
                    min_scale = pd.Timestamp(self.min_value)
                    max_scale = pd.Timestamp(self.max_value)
                    min_converted = (
                        min_scale - pd.Timestamp("1899-12-30")
                    ).total_seconds() / 86400
                    max_converted = (
                        max_scale - pd.Timestamp("1899-12-30")
                    ).total_seconds() / 86400
                    self.chart.x_axis.scaling.min = min_converted
                    self.chart.x_axis.scaling.max = max_converted

                    if self.option["rotronic"] == True:
                        col_time = self.sheet.min_column
                        col_meas_range_min = 2
                        col_meas_range_max = 4
                        while col_time <= self.sheet.max_column:
                            data = Reference(
                                self.sheet,
                                min_col=col_time,
                                min_row=2,
                                max_row=self.sheet.max_row,
                            )
                            col_time += 3
                            for i in range(col_meas_range_min, col_meas_range_max):
                                values = Reference(
                                    self.sheet,
                                    min_col=i,
                                    min_row=1,
                                    max_row=self.sheet.max_row,
                                )
                                series = Series(values, data, title_from_data=True)
                                self.chart.series.append(series)
                            col_meas_range_min += 3
                            col_meas_range_max += 3
                        self.sheet_chart.add_chart(self.chart)
                        for serie in self.chart.series:
                            serie.graphicalProperties.line.width = 15000
                    else:
                        if self.option["indigo"] == True:
                            data = Reference(
                                self.sheet,
                                min_col=1,
                                min_row=2,
                                max_row=self.sheet.max_row,
                            )
                            for i in range(2, 4):
                                values = Reference(
                                    self.sheet,
                                    min_col=i,
                                    min_row=1,
                                    max_row=self.sheet.max_row,
                                )
                                series = Series(values, data, title_from_data=True)
                                self.chart.series.append(series)
                            try:
                                data2 = Reference(
                                    self.sheet,
                                    min_col=4,
                                    min_row=2,
                                    max_row=self.sheet.max_row,
                                )
                                for i in range(5, 7):
                                    values2 = Reference(
                                        self.sheet,
                                        min_col=i,
                                        min_row=1,
                                        max_row=self.sheet.max_row,
                                    )
                                    series2 = Series(
                                        values2, data2, title_from_data=True
                                    )
                                    self.chart.series.append(series2)
                            except:
                                pass

                            self.sheet_chart.add_chart(self.chart)
                            for serie in self.chart.series:
                                serie.graphicalProperties.line.width = 15000
                        else:
                            if self.option["indexZUT"] == True:
                                index = int(6)
                            elif self.option["coolingSys"] == True:
                                index = int(5)
                            else:
                                index = int(self.sheet.max_column + 1)
                            data = Reference(
                                self.sheet,
                                min_col=1,
                                min_row=2,
                                max_row=self.sheet.max_row,
                            )
                            for i in range(2, index):
                                values = Reference(
                                    self.sheet,
                                    min_col=i,
                                    min_row=1,
                                    max_row=self.sheet.max_row,
                                )
                                series = Series(values, data, title_from_data=True)
                                self.chart.series.append(series)
                            self.sheet_chart.add_chart(self.chart)
                            for serie in self.chart.series:
                                serie.graphicalProperties.line.width = 15000

                            if self.option["grafana"] == True:
                                pass
                            else:
                                if self.option["insight"] == True:
                                    # Temperature
                                    s1 = self.chart.series[0]
                                    s1.graphicalProperties.line.solidFill = "F80E0A"

                                    # Humidity
                                    s2 = self.chart.series[0]
                                    s2.graphicalProperties.line.solidFill = "1F95F8"
                                else:
                                    # Temperature Set
                                    s1 = self.chart.series[0]
                                    s1.graphicalProperties.line.solidFill = "752524"

                                    # Temperature
                                    s2 = self.chart.series[1]
                                    s2.graphicalProperties.line.solidFill = "F80E0A"

                                    if self.option["gasChamber"] == True:
                                        # Temperature Set
                                        s1 = self.chart.series[0]
                                        s1.graphicalProperties.line.solidFill = "752524"

                                        # Temperature IN
                                        s2 = self.chart.series[1]
                                        s2.graphicalProperties.line.solidFill = "F80E0A"

                                        # Temperature OUT
                                        s3 = self.chart.series[2]
                                        s3.graphicalProperties.line.solidFill = "F49D1E"

                                        # Humidity Set
                                        s4 = self.chart.series[3]
                                        s4.graphicalProperties.line.solidFill = "173B96"

                                        # Humidity IN
                                        s5 = self.chart.series[4]
                                        s5.graphicalProperties.line.solidFill = "1F95F8"

                                        # Humidity OUT
                                        s6 = self.chart.series[5]
                                        s6.graphicalProperties.line.solidFill = "0AB1C8"
                                    else:
                                        if self.option["saltHum"] == True:
                                            # Humidifier Set
                                            s3 = self.chart.series[2]
                                            s3.graphicalProperties.line.solidFill = (
                                                "0A6E44"
                                            )

                                            # Humidifier
                                            s4 = self.chart.series[3]
                                            s4.graphicalProperties.line.solidFill = (
                                                "2AA31E"
                                            )

                                            # Humidity Set
                                            s5 = self.chart.series[4]
                                            s5.graphicalProperties.line.solidFill = (
                                                "173B96"
                                            )

                                            # Humidity
                                            s6 = self.chart.series[5]
                                            s6.graphicalProperties.line.solidFill = (
                                                "1F95F8"
                                            )
                                        else:
                                            try:
                                                # Humidity Set
                                                s3 = self.chart.series[2]
                                                s3.graphicalProperties.line.solidFill = (
                                                    "173B96"
                                                )
                                            except:
                                                pass

                                            try:
                                                # Humidity
                                                s4 = self.chart.series[3]
                                                s4.graphicalProperties.line.solidFill = (
                                                    "1F95F8"
                                                )
                                            except:
                                                pass

                                            try:
                                                # Temperature Bath
                                                s5 = self.chart.series[4]
                                                s5.graphicalProperties.line.solidFill = (
                                                    "2AA31E"
                                                )
                                            except:
                                                pass

                    self.wb.save(self.savepath)
                    self.popup_win_2.destroy()
                    try:
                        os.startfile(self.savepath)
                    except:
                        Message(
                            "warning",
                            "File can not be opened by Stella"
                            "\nPlease open it manually",
                            0,
                        )

                else:
                    self.sheet_chart = self.wb.create_chartsheet("ChartAll")
                    self.chart = ScatterChart()
                    if self.option["timeabsOne"] == True:
                        self.chart.x_axis.tickLblPos = "low"
                        self.chart.legend.position = "b"

                        self.chart.x_axis.title = "Samples"
                        self.chart.y_axis.title = "Temperature[°C]"

                        # Text format - Axis
                        normal_font = Font(typeface="Calibri")
                        cp_text = CharacterProperties(
                            latin=normal_font, sz=900, b=False
                        )
                        self.chart.x_axis.txPr = RichText(
                            p=[
                                Paragraph(
                                    pPr=ParagraphProperties(defRPr=cp_text),
                                    endParaRPr=cp_text,
                                )
                            ]
                        )
                        self.chart.y_axis.txPr = RichText(
                            p=[
                                Paragraph(
                                    pPr=ParagraphProperties(defRPr=cp_text),
                                    endParaRPr=cp_text,
                                )
                            ]
                        )

                        # Text format - Title
                        cp_text1 = CharacterProperties(
                            latin=normal_font, sz=1000, b=False
                        )
                        pp1 = ParagraphProperties(defRPr=cp_text1)
                        self.chart.x_axis.title.tx.rich.p[0].pPr = pp1
                        self.chart.y_axis.title.tx.rich.p[0].pPr = pp1
                        self.chart.x_axis.txPr.properties.rot = "-2700000"
                        self.chart.y_axis.majorUnit = self.ax_scale.get()
                        min_value = self.sheet["A" + str(self.sheet.min_row + 1)].value
                        max_value = self.sheet["A" + str(self.sheet.max_row)].value
                        self.chart.x_axis.scaling.min = min_value
                        self.chart.x_axis.scaling.max = max_value

                        index = int(self.sheet.max_column + 1)

                        data = Reference(
                            self.sheet, min_col=2, min_row=2, max_row=self.sheet.max_row
                        )
                        for i in range(3, index):
                            values = Reference(
                                self.sheet,
                                min_col=i,
                                min_row=1,
                                max_row=self.sheet.max_row,
                            )
                            series = Series(values, data, title_from_data=True)
                            self.chart.series.append(series)
                        self.sheet_chart.add_chart(self.chart)
                        for serie in self.chart.series:
                            serie.graphicalProperties.line.width = 15000
                    else:
                        if self.option["samples"] == True:
                            self.chart.x_axis.tickLblPos = "low"
                            self.chart.legend.position = "b"

                            self.chart.x_axis.title = "Samples"
                            self.chart.y_axis.title = "Temperature[°C]"

                            # Text format - Axis
                            normal_font = Font(typeface="Calibri")
                            cp_text = CharacterProperties(
                                latin=normal_font, sz=900, b=False
                            )
                            self.chart.x_axis.txPr = RichText(
                                p=[
                                    Paragraph(
                                        pPr=ParagraphProperties(defRPr=cp_text),
                                        endParaRPr=cp_text,
                                    )
                                ]
                            )
                            self.chart.y_axis.txPr = RichText(
                                p=[
                                    Paragraph(
                                        pPr=ParagraphProperties(defRPr=cp_text),
                                        endParaRPr=cp_text,
                                    )
                                ]
                            )

                            # Text format - Title
                            cp_text1 = CharacterProperties(
                                latin=normal_font, sz=1000, b=False
                            )
                            pp1 = ParagraphProperties(defRPr=cp_text1)
                            self.chart.x_axis.title.tx.rich.p[0].pPr = pp1
                            self.chart.y_axis.title.tx.rich.p[0].pPr = pp1
                            self.chart.x_axis.txPr.properties.rot = "-2700000"
                            self.chart.y_axis.majorUnit = self.ax_scale.get()
                            min_value = self.sheet[
                                "A" + str(self.sheet.min_row + 1)
                            ].value
                            max_value = self.sheet["A" + str(self.sheet.max_row)].value
                            self.chart.x_axis.scaling.min = min_value
                            self.chart.x_axis.scaling.max = max_value

                            index = int(self.sheet.max_column + 1)

                            data = Reference(
                                self.sheet,
                                min_col=1,
                                min_row=2,
                                max_row=self.sheet.max_row,
                            )
                            for i in range(3, index):
                                values = Reference(
                                    self.sheet,
                                    min_col=i,
                                    min_row=1,
                                    max_row=self.sheet.max_row,
                                )
                                series = Series(values, data, title_from_data=True)
                                self.chart.series.append(series)
                            self.sheet_chart.add_chart(self.chart)
                            for serie in self.chart.series:
                                serie.graphicalProperties.line.width = 15000
                        else:
                            if self.option["timeabsAll"] == True:
                                self.chart.x_axis.tickLblPos = "low"
                                self.chart.legend.position = "b"

                                self.chart.x_axis.title = "Samples"
                                self.chart.y_axis.title = "Temperature[°C]"

                                # Text format - Axis
                                normal_font = Font(typeface="Calibri")
                                cp_text = CharacterProperties(
                                    latin=normal_font, sz=900, b=False
                                )
                                self.chart.x_axis.txPr = RichText(
                                    p=[
                                        Paragraph(
                                            pPr=ParagraphProperties(defRPr=cp_text),
                                            endParaRPr=cp_text,
                                        )
                                    ]
                                )
                                self.chart.y_axis.txPr = RichText(
                                    p=[
                                        Paragraph(
                                            pPr=ParagraphProperties(defRPr=cp_text),
                                            endParaRPr=cp_text,
                                        )
                                    ]
                                )

                                # Text format - Title
                                cp_text1 = CharacterProperties(
                                    latin=normal_font, sz=1000, b=False
                                )
                                pp1 = ParagraphProperties(defRPr=cp_text1)
                                self.chart.x_axis.title.tx.rich.p[0].pPr = pp1
                                self.chart.y_axis.title.tx.rich.p[0].pPr = pp1
                                self.chart.x_axis.txPr.properties.rot = "-2700000"
                                self.chart.y_axis.majorUnit = self.ax_scale.get()
                                min_value = self.sheet[
                                    "A" + str(self.sheet.min_row + 1)
                                ].value
                                max_value = self.sheet[
                                    "A" + str(self.sheet.max_row)
                                ].value
                                self.chart.x_axis.scaling.min = min_value
                                self.chart.x_axis.scaling.max = max_value

                                col_time = 3
                                col_meas = 2
                                while col_time <= self.sheet.max_column:
                                    data = Reference(
                                        self.sheet,
                                        min_col=col_time,
                                        min_row=2,
                                        max_row=self.sheet.max_row,
                                    )
                                    col_time += 2
                                    if col_meas <= self.sheet.max_column:
                                        values = Reference(
                                            self.sheet,
                                            min_col=col_meas,
                                            min_row=1,
                                            max_row=self.sheet.max_row,
                                        )
                                        series = Series(
                                            values, data, title_from_data=True
                                        )
                                        self.chart.series.append(series)
                                    col_meas += 2
                                self.sheet_chart.add_chart(self.chart)
                                for serie in self.chart.series:
                                    serie.graphicalProperties.line.width = 15000
                            else:
                                if self.option["secasi"] == True:
                                    self.chart.x_axis.tickLblPos = "low"
                                    self.chart.legend.position = "b"

                                    self.chart.x_axis.title = "Time [s]"
                                    self.chart.y_axis.title = "Temperature[°C]"

                                    # Text format - Axis
                                    normal_font = Font(typeface="Calibri")
                                    cp_text = CharacterProperties(
                                        latin=normal_font, sz=900, b=False
                                    )
                                    self.chart.x_axis.txPr = RichText(
                                        p=[
                                            Paragraph(
                                                pPr=ParagraphProperties(defRPr=cp_text),
                                                endParaRPr=cp_text,
                                            )
                                        ]
                                    )
                                    self.chart.y_axis.txPr = RichText(
                                        p=[
                                            Paragraph(
                                                pPr=ParagraphProperties(defRPr=cp_text),
                                                endParaRPr=cp_text,
                                            )
                                        ]
                                    )

                                    # Text format - Title
                                    cp_text1 = CharacterProperties(
                                        latin=normal_font, sz=1000, b=False
                                    )
                                    pp1 = ParagraphProperties(defRPr=cp_text1)
                                    self.chart.x_axis.title.tx.rich.p[0].pPr = pp1
                                    self.chart.y_axis.title.tx.rich.p[0].pPr = pp1
                                    self.chart.x_axis.txPr.properties.rot = "-2700000"
                                    self.chart.y_axis.majorUnit = self.ax_scale.get()
                                    min_value = self.sheet[
                                        "A" + str(self.sheet.min_row + 1)
                                    ].value
                                    max_value = self.sheet[
                                        "A" + str(self.sheet.max_row)
                                    ].value
                                    self.chart.x_axis.scaling.min = min_value
                                    self.chart.x_axis.scaling.max = max_value
                                else:
                                    self.chart.x_axis.number_format = (
                                        "yyyy-mm-dd HH:MM:SS"
                                    )
                                    self.chart.x_axis.majorTimeUnit = "months"
                                    self.chart.x_axis.tickLblPos = "low"
                                    self.chart.legend.position = "b"
                                    self.chart.x_axis.title = (
                                        "Date and Time [yyyy-mm-dd hh:mm:ss]"
                                    )
                                    self.chart.y_axis.title = "Temperature[°C]"

                                    # Text format - Axis
                                    normal_font = Font(typeface="Calibri")
                                    cp_text = CharacterProperties(
                                        latin=normal_font, sz=900, b=False
                                    )
                                    self.chart.x_axis.txPr = RichText(
                                        p=[
                                            Paragraph(
                                                pPr=ParagraphProperties(defRPr=cp_text),
                                                endParaRPr=cp_text,
                                            )
                                        ]
                                    )
                                    self.chart.y_axis.txPr = RichText(
                                        p=[
                                            Paragraph(
                                                pPr=ParagraphProperties(defRPr=cp_text),
                                                endParaRPr=cp_text,
                                            )
                                        ]
                                    )

                                    # Text format - Title
                                    cp_text1 = CharacterProperties(
                                        latin=normal_font, sz=1000, b=False
                                    )
                                    pp1 = ParagraphProperties(defRPr=cp_text1)
                                    self.chart.x_axis.title.tx.rich.p[0].pPr = pp1
                                    self.chart.y_axis.title.tx.rich.p[0].pPr = pp1
                                    self.chart.x_axis.txPr.properties.rot = "-2700000"
                                    self.chart.y_axis.majorUnit = self.ax_scale.get()
                                    min_value = self.sheet[
                                        "A" + str(self.sheet.min_row + 1)
                                    ].value
                                    max_value = self.sheet[
                                        "A" + str(self.sheet.max_row)
                                    ].value
                                    min_scale = pd.Timestamp(min_value)
                                    max_scale = pd.Timestamp(max_value)
                                    min_converted = (
                                        min_scale - pd.Timestamp("1899-12-30")
                                    ).total_seconds() / 86400
                                    max_converted = (
                                        max_scale - pd.Timestamp("1899-12-30")
                                    ).total_seconds() / 86400
                                    self.chart.x_axis.scaling.min = min_converted
                                    self.chart.x_axis.scaling.max = max_converted
                                if self.option["keithleymanager"] == True:
                                    col_time = 1
                                    col_meas = 2
                                    while col_time <= self.sheet.max_column:
                                        data = Reference(
                                            self.sheet,
                                            min_col=col_time,
                                            min_row=2,
                                            max_row=self.sheet.max_row,
                                        )
                                        col_time += 2
                                        if col_meas <= self.sheet.max_column:
                                            values = Reference(
                                                self.sheet,
                                                min_col=col_meas,
                                                min_row=1,
                                                max_row=self.sheet.max_row,
                                            )
                                            series = Series(
                                                values, data, title_from_data=True
                                            )
                                            self.chart.series.append(series)
                                        col_meas += 2
                                    self.sheet_chart.add_chart(self.chart)
                                    for serie in self.chart.series:
                                        serie.graphicalProperties.line.width = 15000
                                else:
                                    if self.option["indexZUT"] == True:
                                        index = int(6)
                                    elif self.option["coolingSys"] == True:
                                        index = int(5)

                                    else:
                                        index = int(self.sheet.max_column + 1)
                                    data = Reference(
                                        self.sheet,
                                        min_col=1,
                                        min_row=2,
                                        max_row=self.sheet.max_row,
                                    )
                                    for i in range(2, index):
                                        values = Reference(
                                            self.sheet,
                                            min_col=i,
                                            min_row=1,
                                            max_row=self.sheet.max_row,
                                        )
                                        series = Series(
                                            values, data, title_from_data=True
                                        )
                                        self.chart.series.append(series)
                                    self.sheet_chart.add_chart(self.chart)
                                    for serie in self.chart.series:
                                        serie.graphicalProperties.line.width = 15000
                                if (
                                    self.option["agilent"] == True
                                    or self.option["keithleymanager"] == True
                                ):
                                    pass
                                else:
                                    if self.option["insight"] == True:
                                        # Temperature
                                        s1 = self.chart.series[0]
                                        s1.graphicalProperties.line.solidFill = "F80E0A"
                                    else:
                                        # Temperature Hot Set
                                        s2 = self.chart.series[0]
                                        s2.graphicalProperties.line.solidFill = "752524"

                                        # Temperature Hot
                                        s3 = self.chart.series[1]
                                        s3.graphicalProperties.line.solidFill = "F80E0A"

                                        if self.option["humBath"] == True:
                                            # Humidifier Set
                                            s4 = self.chart.series[2]
                                            s4.graphicalProperties.line.solidFill = (
                                                "0A6E44"
                                            )

                                            # Humidifier
                                            s5 = self.chart.series[3]
                                            s5.graphicalProperties.line.solidFill = (
                                                "2AA31E"
                                            )
                                        else:
                                            try:
                                                # Temperature Cold Set
                                                s6 = self.chart.series[2]
                                                s6.graphicalProperties.line.solidFill = (
                                                    "173B96"
                                                )
                                            except:
                                                pass
                                            try:
                                                # Temperature Cold
                                                s7 = self.chart.series[3]
                                                s7.graphicalProperties.line.solidFill = (
                                                    "1F95F8"
                                                )
                                            except:
                                                pass
                                            try:
                                                # Temperature Basket
                                                s8 = self.chart.series[4]
                                                s8.graphicalProperties.line.solidFill = (
                                                    "2AA31E"
                                                )
                                            except:
                                                pass

                    self.wb.save(self.savepath)
                    self.popup_win_2.destroy()
                    try:
                        os.startfile(self.savepath)
                    except:
                        Message(
                            "warning",
                            "File can not be opened by Stella"
                            "\nPlease open it manually",
                            0,
                        )
            except:
                try:
                    self.popup_win_2.destroy()
                except:
                    pass
                Message(
                    "error",
                    "Something went wrong"
                    "\nPlease contact with: piotr.klys92@gmail.com",
                    0,
                )
